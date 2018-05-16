
from datetime import date
from datetime import datetime
import psycopg2
from openpyxl.reader.excel import load_workbook
def insert_data_into_postgres(table_name,path):
    conn = psycopg2.connect(database='dbname', user = 'postgres', password = 'balaji1212', host = "127.0.0.1", port = "5432")
    
    
    print("Opened database successfully")
    cur = conn.cursor()
    cur2= conn.cursor()
    #path = "F:\Mongo\\test.xlsx"
    wb = load_workbook(filename=path)
    ws=wb.get_sheet_by_name("Sheet1")
    print(("Inserting data into {}").format(table_name))
    for row_val in range(2,ws.max_row+1):
        #print("in For Loop")
        el=[]
        #s=(ws.cell(row=row_val, column = 1).value)
        for coulmn_val in range(1,(ws.max_column)+1):
            #s=s+","+(ws.cell(row=row_val, column = coulmn_val).value)
            el.append((ws.cell(row=row_val, column = coulmn_val).value))
        #print(el)
        #print(("INSERT INTO {} VALUES('{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}')").format(table_name,el[0],el[1],datetime.date(el[2]),el[3],el[4],el[5],el[6],el[7],el[8],el[9],el[10]))
        cur.execute(("INSERT INTO {} VALUES('{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}')").format(table_name,el[0],el[1],datetime.date(el[2]),el[3],el[4],el[5],el[6],el[7],el[8],el[9],el[10]))
        #print("done")
        conn.commit()
    print(ws.max_row,(" records inserted into {} successfully").format(table_name))
    conn.close()

#insert_data_into_postgres('superstore', "F:\Mongo\\test.xlsx")